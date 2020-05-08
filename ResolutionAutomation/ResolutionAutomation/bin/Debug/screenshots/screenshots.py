from openpyxl.drawing.image import Image
import openpyxl
import os

# 目录路径
currentPath = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.Workbook()
ws = wb.worksheets[0]

img = Image(currentPath+"/1024×768.png")
ws['A2'] = '1024×768'
ws.add_image(img, "B2")

img = Image(currentPath+"/1152×864.png")
ws['A46'] = '1152×864'
ws.add_image(img, "B46")

img = Image(currentPath+"/1280×600.png")
ws['A95'] = '1280×600'
ws.add_image(img, "B95")

img = Image(currentPath+"/1360×768.png")
ws['A130'] = '1360×768'
ws.add_image(img, "B130")

img = Image(currentPath+"/1366×768.png")
ws['A174'] = '1366×768'
ws.add_image(img, "B174")

img = Image(currentPath+"/1400×1050.png")
ws['A218'] = '1400×1050'
ws.add_image(img, "B218")

img = Image(currentPath+"/1440×900.png")
ws['A278'] = '1440×900'
ws.add_image(img, "B278")

img = Image(currentPath+"/1600×900.png")
ws['A329'] = '1600×900'
ws.add_image(img, "B329")

img = Image(currentPath+"/1680×1050.png")
ws['A380'] = '1680×1050'
ws.add_image(img, "B380")

img = Image(currentPath+"/1920×1080.png")
ws['A440'] = '1920×1080'
ws.add_image(img, "B440")

wb.save(currentPath+"/screenshots.xlsx")