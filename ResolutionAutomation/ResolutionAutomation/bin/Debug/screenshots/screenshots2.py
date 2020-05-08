from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font
import openpyxl
import os

# 目录路径
currentPath = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.Workbook()

# 创建
dirs2 = os.listdir("../script")
for file in dirs2:
	script = file[0:-3]
	ws = wb.create_sheet(script)
wb.save(currentPath+"/screenshots2.xlsx")

# 修改
wb2 = openpyxl.load_workbook(currentPath+"/screenshots2.xlsx")
for file in dirs2:
	script = file[0:-3]
	# print(script)
	
	ws2 = wb2[script]
	i = 2
	for root, dirs, files in os.walk("."):
		dirs.sort()
		for name in dirs:
			# print(name)

			img = Image(currentPath+"/"+name+"/"+script+".png")
			ws2["A%d"%(i)] = name
			ws2["A%d"%(i)].font = Font(name='Yu Gothic',size=22,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
			ws2.add_image(img, "B%d"%(i+1))

			if script == "1024×768":
				i += (32 + 13)
			elif script == "1152×864":
				i += (36 + 14)
			elif script == "1280×600":
				i += (26 + 9)
			elif script == "1360×768":
				i += (32 + 13)
			elif script == "1366×768":
				i += (32 + 13)
			elif script == "1400×1050":
				i += (43 + 18)
			elif script == "1440×900":
				i += (37 + 16)
			elif script == "1600×900":
				i += (37 + 16)
			elif script == "1680×1050":
				i += (43 + 18)
			elif script == "1920×1080":
				i += (44 + 19)

			# if script == "1024×768":
			# 	i += 32
			# elif script == "1152×864":
			# 	i += 36
			# elif script == "1280×600":
			# 	i += 26
			# elif script == "1360×768":
			# 	i += 32
			# elif script == "1366×768":
			# 	i += 32
			# elif script == "1400×1050":
			# 	i += 43
			# elif script == "1440×900":
			# 	i += 37
			# elif script == "1600×900":
			# 	i += 37
			# elif script == "1680×1050":
			# 	i += 43
			# elif script == "1920×1080":
			# 	i += 44

	
wb2.save(currentPath+"/screenshots2.xlsx")